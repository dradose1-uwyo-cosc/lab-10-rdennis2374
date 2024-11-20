# Ryan Dennis
# UWYO COSC 1010
# Submission Date 11/18/2024
# Homework 5
# Lab Section: 14
# Sources, people worked with, help given to: Stack Overflow, Geeks for Geeks
# your
# comments
# here
# refrence image from smartkidworld.com
from pathlib import Path
import openpyxl
from openpyxl.styles import Color,PatternFill
import string
wb = openpyxl.Workbook()
sheet = wb.active
letters = ['A','B','C','D','E','F','G','H','I','J']
for i in range (1,12):
    sheet.row_dimensions[i].height = 29
for letter in letters:
    sheet.column_dimensions[letter].width = 5
orange_fill = PatternFill(patternType='solid', fgColor='FFA500')
yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
red_fill = PatternFill(patternType='solid', fgColor='FF0000')
dark_green_fill = PatternFill(patternType= 'solid', fgColor='013220')
light_green_fill = PatternFill(patternType= 'solid', fgColor='008000')
Red = ["C4","G4","C5","D5","F5","G5","D6","E6","F6"]
Light_Green = ["I5","B6","H6","I6","B7","C7","G7","H7","C8","D9"]
Orange = ["C1","G1","B2","C2","G2","H2","B3","C3","D3","E3","F3","G3","H3","D4","E4","F4","E5"]
Dark_Green = ["E7","E8","G8","C9","E9","F9","G9","D10","E10","F10"]
Yellow = ["E1","E2"]
for cell in Red:
    sheet[cell].fill = red_fill
for cell in Light_Green:
    sheet[cell].fill = light_green_fill
for cell in Orange:
    sheet[cell].fill = orange_fill
for cell in Dark_Green:
    sheet[cell].fill = dark_green_fill
for cell in Yellow:
    sheet[cell].fill = yellow_fill
wb.save("art.xlsx")